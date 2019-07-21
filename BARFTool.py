import tkinter
from tkinter import Tk, ttk
from openpyxl import load_workbook

wb = load_workbook('Master Org Chart_Technical Services.xlsx')
bmt_sheets = ['BM Capitol Lakes', 'BM SE', 'BM MidAtlantic', 'BM Gulfcoast', 'BM NE', 'BM Great Plains', 'BM Midwest',
              'BM Western']
atom_sheets = ['Capitol Lakes', 'Southeast', 'Mid Atlantic', 'Gulf Coast', 'Northeast', 'Great Plains', 'Midwest',
              'Western']

def get_results(a):
    personnel = []
    records = resultsTree.get_children()
    for element in records:
        resultsTree.delete(element)
    clinic = int(entryBox.get())
    clinicChoice.set(clinic)
    entryBox.delete(0, 'end')
    personnel = find_bmts(clinic, personnel)
    atom_row, personnel = find_atoms(clinic, personnel)
    personnel = find_atoms_2(clinic, personnel)
    rtom_email, personnel = find_rtom(atom_row, personnel)
    personnel = find_fvp(rtom_email, personnel)
    for item in personnel:
        resultsTree.insert('', 'end', text=item[0], values=(item[1], item[2], item[3]))



def find_bmts(clinic, personnel):
    for sheet in bmt_sheets:
        ws = wb[sheet]
        bmt = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=ws.max_column):
            for cell in row:
                if cell.value == clinic or cell.value == "#" + str(clinic):
                    if ws.cell(row=cell.row, column=3).value != "ATOM":
                        bmt.append(ws.cell(row=cell.row, column=2).value)
                        bmt.append(ws.cell(row=cell.row, column=3).value)
                        bmt.append(ws.cell(row=cell.row, column=8).value)
                        bmt.append(ws.cell(row=cell.row, column=9).value)
                        personnel.append(bmt)
                        bmt = []
    return personnel


def find_atoms_2(clinic, personnel):
    for sheet in atom_sheets:
        ws = wb[sheet]
        atom = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=15, max_col=ws.max_column):
            for cell in row:
                if cell.value == clinic or cell.value == "#" + str(clinic):
                    atom.append(ws.cell(row=cell.row, column=2).value)
                    atom.append(ws.cell(row=cell.row, column=3).value)
                    atom.append(ws.cell(row=cell.row, column=8).value)
                    atom.append(ws.cell(row=cell.row, column=9).value)
                    if atom in personnel:
                        print("duplicate ATOM")
                    else:
                        personnel.append(atom)
                    atom = []
    return personnel

def find_atoms(clinic, personnel):
    ws = wb['ATOMS']
    atom_row = 0
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=15, max_col=31):
        atom = []
        for cell in row:
            if str(clinic) in str(cell.value):  #  or "#" + str(clinic) in str(cell.value):
                atom.append(ws.cell(row=cell.row, column=2).value)
                atom.append(ws.cell(row=cell.row, column=3).value)
                atom.append(ws.cell(row=cell.row, column=8).value)
                atom.append(ws.cell(row=cell.row, column=9).value)
                personnel.append(atom)
                atom_row = cell.row
    return atom_row, personnel


def find_rtom(atom_row, personnel):
    ws = wb['ATOMS']
    rtom_email = ""
    rtom = []
    for i in range(atom_row, 1, -1):
        if ws.cell(row=i, column=3).value == "RTOM":
            rtom_email = ws.cell(row=i, column=9).value
            rtom.append(ws.cell(row=i, column=2).value)
            rtom.append(ws.cell(row=i, column=3).value)
            rtom.append(ws.cell(row=i, column=8).value)
            rtom.append(ws.cell(row=i, column=9).value)
            personnel.append(rtom)
            break
    return rtom_email, personnel


def find_fvp(rtom_email, personnel):
    ws = wb['RTOMS']
    fvp = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=10, max_col=10):
        for cell in row:
            if cell.value == rtom_email:
                for i in range(cell.row, 1, -1):
                    if ws.cell(row=i, column=3).value == "Field Vice President":
                        fvp.append(ws.cell(row=i, column=2).value)
                        fvp.append(ws.cell(row=i, column=3).value)
                        fvp.append(ws.cell(row=i, column=9).value)
                        fvp.append(ws.cell(row=i, column=10).value)
                        personnel.append(fvp)
                        break
    return personnel


def select_barf(a):
    try:
        selection = resultsTree.item(resultsTree.selection()).get("values")[2]
        r = Tk()
        r.clipboard_clear()
        r.clipboard_append(selection)
        r.update()
        r.destroy()
    except IndexError:
        pass


mainWindow = tkinter.Tk()
mainWindow.title("BARFTool 1.0")
mainWindow.option_add("*Font", "arial 9")
mainWindow.option_add("*Background", "#F0F0F0")
style = ttk.Style()
style.configure('Treeview.Heading', foreground='black', bg='#F0F0F0', font=('arial', 9))
clinicChoice = tkinter.Variable(mainWindow)

entryFrame = tkinter.Frame(mainWindow)
entryFrame.grid(row=0, column=0)
entryLabel = tkinter.Label(entryFrame, text="Enter clinic number: ")
entryLabel.grid(row=0, column=0)
entryBox = tkinter.Entry(entryFrame)
entryBox.grid(row=0, column=1)
choiceLabel = tkinter.Label(entryFrame, textvariable=clinicChoice)
choiceLabel.grid(row=0, column=2)

resultsFrame = tkinter.Frame(mainWindow)
resultsFrame.grid(row=1, column=0)
resultsTree = ttk.Treeview(resultsFrame, height=12, column=['', '', ''], style='Treeview.Heading')
resultsTree.column('#0', width=200)
resultsTree.column('#1', width=200)
resultsTree.column('#2', width=100)
resultsTree.column('#3', width=200)
resultsTree.grid(row=0, column=0)
resultsTree.heading('#0', text='Name')
resultsTree.heading('#1', text='Position')
resultsTree.heading('#2', text='Phone')
resultsTree.heading('#3', text='Email')

entryBox.focus()
entryBox.bind("<Return>", get_results)
resultsTree.bind('<ButtonRelease-1>', select_barf)

mainWindow.mainloop()
